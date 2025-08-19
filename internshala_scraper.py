import requests
from bs4 import BeautifulSoup
import pandas as pd
import time
from openpyxl import load_workbook
from openpyxl.styles import Alignment,Font

# Target URL
url = "https://internshala.com/jobs/python-django-jobs/"
headers = {
    "User-Agent": "Mozilla/5.0"
}

# Fetch HTML
response = requests.get(url, headers=headers)
soup = BeautifulSoup(response.text, "html.parser")

# Find job cards
job_cards = soup.find_all("div", class_="internship_meta experience_meta")

jobs = []

# Extract job details
for card in job_cards:
    try:
        title = card.find("a", class_="job-title-href").text.strip()
    except:
        title = ""

    loc_icon = card.find("i", class_=["ic-16-map-pin", "ic-16-home"])
    if loc_icon:
        location = loc_icon.find_next("span").text.strip()
    else:
        location = ""

    try:
        job_url = "https://internshala.com" + card.find("a", class_="job-title-href")["href"]
    except:
        job_url = ""    


    skills = ""
    job_desc = ""
    if job_url:
        try:
            detail_resp = requests.get(job_url, headers=headers)
            detail_soup = BeautifulSoup(detail_resp.text, "html.parser")
            # Skills
            skills_div = detail_soup.find("div", class_="round_tabs_container")
            if skills_div:
                skill_spans = skills_div.find_all("span", class_="round_tabs")
                skills = ", ".join([s.text.strip() for s in skill_spans])
            # Job Description
            internship_details = detail_soup.find("div", class_="internship_details")
            if internship_details:
                text_container = internship_details.find("div", class_="text-container")
                if text_container:
                    lines = [line.strip() for line in text_container.text.split('\n') if line.strip()]
                    if lines:
                        job_desc = lines[0:10]
        except Exception as e:
            skills = ""
            job_desc = ""
        time.sleep(0.5)     
    
    
    brief_icon = card.find("i", class_=["ic-16-briefcase"])
    if brief_icon:
        experience = brief_icon.find_next("span").text.strip()
    else:
        experience = ""    


    try:
        salary = card.find("span", class_="desktop").text.strip()
    except:
        salary = ""


    jobs.append({
        "JobTitle": title,
        "Location": location,
        "ExperienceRequired": experience,
        "SkillsRequired": skills,
        "Salary": salary,
        "JobDescriptionSummary": job_desc,
        "JobURL": job_url 
    })

# Save to Excel
df = pd.DataFrame(jobs)
df.to_excel("Internshala_Jobs.xlsx", index=False, engine="openpyxl")

wb = load_workbook("Internshala_Jobs.xlsx")
ws = wb.active

# Auto-adjust column widths and set wrap_text for all cells
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        # Set wrap_text for all cells
        cell.alignment = Alignment(wrap_text=True)
        # Find maximum length in this column
        try:
            cell_length = len(str(cell.value))
            if cell_length > max_length:
                max_length = cell_length
        except:
            pass
    ws.column_dimensions[column_letter].width = max_length + 2

# Make header row bold
for cell in ws[1]:
    cell.font = Font(bold=True)

wb.save("Internshala_Jobs.xlsx")

print("✅ Scraping complete! Saved as Internshala_Jobs.xlsx")
